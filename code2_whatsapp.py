import os
import pyautogui as gui
import time
from openpyxl import load_workbook
import webbrowser
import sys
    
wb = load_workbook("FILE_FINAL_2.xlsx")
sheet = wb['Sheet1']

positionX1, positionY1 = 336, 145  #Search Area
positionX2, positionY2 = 745, 994  #Message Area

message = "Message here"

print('GO!')
#webbrowser.open('https://web.whatsapp.com')
time.sleep(10)

try:
    for audience in range(1, sheet.max_row+1,1):
        gui.click(positionX1, positionY1)
        time.sleep(0.5)
        gui.hotkey('ctrl', 'a')
        time.sleep(0.4)
        gui.write(str(sheet["A"+str(audience)].value), interval=0.025)
        time.sleep(0.8)
        gui.press('enter')
        gui.click(positionX2, positionY2)
        time.sleep(0.5)
        gui.hotkey('ctrl', 'a')
        time.sleep(0.4)
        gui.write(message)
    ##    File = open("message.txt")
    ##    for line in File:
    ##        message = line.rstrip()
    ##        gui.write(message)
    ##        gui.hotkey('shift','enter')
        gui.press('enter')
        print("message sent to "+str(sheet["A"+str(audience)].value))

except KeyboardInterrupt:
    sys.exit()
